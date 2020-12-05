# -*-coding:utf-8-*-
import jwt
import re
import websocket

try:
    import thread
except ImportError:
    import _thread as thread
import json
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime


def get_sign(room_id, app_id, secret):
    payload = {
        "iat": int(time.time()),
        "exp": int(time.time()) + 600,
        "roomId": room_id,
        "appId": app_id
    }
    token = jwt.encode(payload, secret, algorithm='HS256')

    return token


def on_message(ws, message):
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    info = json.loads(message)
    print(info)
    data = message
    # print("数据过滤中...")
    username = re.findall('"badgeName":"(.*?)","content":".*?",', data)
    comment = re.findall('"badgeName":".*?","content":"(.*?)",', data)
    if username and comment:
        d = '{"%s":"%s"}' % (username[0], comment[0])
        comment_data = json.loads(d)
        # print(username,comment)

        for name, comment in comment_data.items():
            wb = load_workbook(FILE_NAME)
            ws0 = wb.active
            ws0.append([name, comment, current_time])
            wb.save(FILE_NAME)


def on_error(ws, error):
    print(error)


def on_close(ws):
    print("### closed ###")


def on_open(ws):
    def run(*args):
        ws.send(
            '{"command":"subscribeNotice","data":["getMessageNotice","getVipEnterBannerNotice","getSendItemNotice","getOnTVAwardNotice", "getOpenNobleNotice", "getOpenGuardianNotice", "getUserMutedNotice"],"reqId":"123456789"}')
        while True:
            ws.send("ping")
            time.sleep(10)

    thread.start_new_thread(run, ())


if __name__ == "__main__":
    room_id = input("请输入直播间链接：").split("/")[-1]
    # room_id = 10778
    app_id = '975ef3083b866918'
    secret = 'adddc8acbe8f7f2204671e0c7294d6e1'
    sign = get_sign(room_id, app_id, secret)
    print(sign.decode())

    FILE_NAME = "虎牙直播间%s.xlsx" % room_id
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "弹幕数据表"
    ws0.append(["用户名", "评论内容", "评论时间"])
    wb.save(FILE_NAME)

    websocket.enableTrace(True)
    ws = websocket.WebSocketApp(
        "ws://ws-apiext.huya.com/index.html?do=comm&roomId=" + str(room_id) + "&appId=" + app_id + "&iat=" + str(
            int(time.time())) + "&sToken=" + sign.decode(),
        on_message=on_message,
        on_error=on_error,
        on_close=on_close)
    ws.on_open = on_open
    ws.run_forever()
